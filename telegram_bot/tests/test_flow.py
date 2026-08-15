from __future__ import annotations

import tempfile
import unittest
from shutil import copy2
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from telegram_bot.app import BotApp
from telegram_bot.config import Config
from telegram_bot.database import Database, PrizeLimitReached
from telegram_bot.workbook import LEADS_HEADERS, PromoWorkbook


PROMO_HEADERS = [
    "Код приза",
    "Уровень",
    "Категория",
    "Приз",
    "Условия применения",
    "Промокод",
    "Статус",
    "Дата использования",
    "Получатель / клиент",
    "Комментарий",
    "№ в пуле",
    "Контроль",
]


class FakeTelegramApi:
    def __init__(self):
        self.messages = []
        self.photos = []
        self.callback_answers = []
        self.member = {"status": "left"}

    def send_message(self, chat_id, text, **kwargs):
        self.messages.append({"chat_id": chat_id, "text": text, **kwargs})
        return {"message_id": len(self.messages)}

    def send_photo(self, chat_id, photo_bytes, filename="welcome_logo.png", caption="", **kwargs):
        payload = {"chat_id": chat_id, "photo_bytes": photo_bytes, "filename": filename, "caption": caption, "text": caption, **kwargs}
        self.photos.append(payload)
        self.messages.append(payload)
        return {"message_id": len(self.messages)}

    def set_my_description(self, description, language_code=""):
        return True

    def set_my_short_description(self, short_description, language_code=""):
        return True

    def answer_callback_query(self, callback_query_id, text="", show_alert=False):
        self.callback_answers.append(
            {"id": callback_query_id, "text": text, "show_alert": show_alert}
        )
        return True

    def get_chat_member(self, chat_id, user_id):
        return self.member


def create_test_workbook(path: Path) -> None:
    workbook = Workbook()
    promo = workbook.active
    promo.title = "Промокоды"
    for column, title in enumerate(PROMO_HEADERS, 1):
        promo.cell(4, column, title)
    promo.append([])
    rows = [
        [
            "TEST-PRIZE",
            "1 уровень",
            "Тест",
            "Тестовый приз",
            "Показать промокод менеджеру",
            "FW-TEST-0001",
            "Не использован",
            None,
            "",
            "",
            1,
            "OK",
        ],
        [
            "TEST-PRIZE",
            "1 уровень",
            "Тест",
            "Тестовый приз",
            "Показать промокод менеджеру",
            "FW-TEST-0002",
            "Не использован",
            None,
            "",
            "",
            2,
            "OK",
        ],
    ]
    for row_index, values in enumerate(rows, 5):
        for column, value in enumerate(values, 1):
            promo.cell(row_index, column, value)

    leads = workbook.create_sheet("Лиды Telegram")
    for column, title in enumerate(LEADS_HEADERS, 1):
        leads.cell(4, column, title)
    workbook.save(path)
    workbook.close()


class BotFlowTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.xlsx = root / "promos.xlsx"
        create_test_workbook(self.xlsx)
        self.config = Config(
            bot_token="test-token",
            bot_username="FlexyTestBot",
            channel_id="@flexy_test",
            channel_url="https://t.me/flexy_test",
            workbook_path=self.xlsx,
            database_path=root / "bot.sqlite3",
            timezone="Europe/Moscow",
            token_ttl_hours=24,
            polling_timeout=1,
            manager_message="Менеджер свяжется с вами.",
            data_secret_path=root / "local_secret.key",
            data_retention_days=30,
            excel_pii_mode="masked",
            subscription_required=True,
            organizer_name="Flexy Way",
            privacy_contact="@flexy_test",
            rate_limit_per_minute=20,
            prize_limit_exempt_username="RRedactor",
        )
        self.workbook = PromoWorkbook(self.xlsx, self.config.timezone)
        self.database = Database(self.config.database_path)
        self.database.initialize()
        self.database.sync_inventory(self.workbook.inventory())
        self.api = FakeTelegramApi()
        self.app = BotApp(self.config, self.database, self.workbook, self.api)
        self.user = {
            "id": 100500,
            "username": "winner",
            "first_name": "Анна",
            "last_name": "Иванова",
        }
        self.claim = self.database.create_token("game-session-1", "TEST-PRIZE", 24)

    def tearDown(self):
        self.database.close()
        self.temp_dir.cleanup()

    def _start(self):
        self.app.handle_update(
            {
                "update_id": 1,
                "message": {
                    "chat": {"id": self.user["id"], "type": "private"},
                    "from": self.user,
                    "text": f"/start {self.claim['token']}",
                },
            }
        )

    def _contact(self, user_id=None):
        self.app.handle_update(
            {
                "update_id": 2,
                "message": {
                    "chat": {"id": self.user["id"], "type": "private"},
                    "from": self.user,
                    "contact": {
                        "user_id": self.user["id"] if user_id is None else user_id,
                        "phone_number": "79991234567",
                    },
                },
            }
        )

    def _consent(self):
        self.app.handle_update(
            {
                "update_id": 10,
                "callback_query": {
                    "id": "consent",
                    "from": self.user,
                    "data": "consent",
                    "message": {"chat": {"id": self.user["id"], "type": "private"}},
                },
            }
        )

    def _check(self, callback_id="cb-1"):
        self.app.handle_update(
            {
                "update_id": 3,
                "callback_query": {
                    "id": callback_id,
                    "from": self.user,
                    "data": "check_subscription",
                    "message": {
                        "chat": {"id": self.user["id"], "type": "private"}
                    },
                },
            }
        )

    def test_full_flow_is_idempotent_and_updates_excel(self):
        self._start()
        self.assertIn("Все призы действуют только для новых клиентов", self.api.messages[-1]["text"])
        self.assertIn("Скидки и бонусы не суммируются", self.api.messages[-1]["text"])
        consent_button = self.api.messages[-1]["reply_markup"]["inline_keyboard"][0][0]
        self.assertEqual(consent_button["callback_data"], "consent")
        self._consent()
        contact_button = self.api.messages[-1]["reply_markup"]["keyboard"][0][0]
        self.assertTrue(contact_button["request_contact"])

        self._contact()
        stored = self.database.get_claim(self.claim["token"])
        self.assertEqual(stored["phone"], "+79991234567")
        self.assertIn("inline_keyboard", self.api.messages[-1]["reply_markup"])

        self._check("not-member")
        self.assertEqual(self.database.get_claim(self.claim["token"])["status"], "pending")
        self.assertIn("Подписка пока не найдена", self.api.messages[-1]["text"])

        self.api.member = {"status": "member"}
        self._check("member")
        issued = self.database.get_claim(self.claim["token"])
        self.assertEqual(issued["status"], "issued")
        self.assertEqual(issued["promo_code"], "FW-TEST-0001")
        self.assertIn("FW-TEST-0001", self.api.messages[-1]["text"])
        self.assertIn("Общие условия получения", self.api.messages[-1]["text"])

        workbook = load_workbook(self.xlsx, read_only=True, data_only=False)
        try:
            promo = workbook["Промокоды"]
            self.assertEqual(promo["G5"].value, "Использован")
            self.assertEqual(promo["F6"].value, "FW-TEST-0002")
            self.assertEqual(promo["G6"].value, "Не использован")
            leads = workbook["Лиды Telegram"]
            self.assertEqual(leads["B5"].value, issued["telegram_user_ref"])
            self.assertTrue(str(leads["C5"].value).endswith("4567"))
            self.assertEqual(leads["I5"].value, "FW-TEST-0001")
            self.assertEqual(leads["J5"].value, issued["token_ref"])
            self.assertEqual(leads["L5"].value, "Выдан")
        finally:
            workbook.close()

        self._check("repeat")
        summary = self.database.prize_summary()[0]
        self.assertEqual(summary["available"], 1)
        self.assertEqual(self.database.get_claim(self.claim["token"])["promo_code"], "FW-TEST-0001")

    def test_rejects_foreign_contact(self):
        self._start()
        self._consent()
        self._contact(user_id=777)
        self.assertEqual(self.database.get_claim(self.claim["token"])["phone"], "")
        self.assertIn("именно свой номер", self.api.messages[-1]["text"])

    def test_token_is_locked_to_first_telegram_user(self):
        self._start()
        other = {"id": 42, "first_name": "Другой"}
        self.app.handle_update(
            {
                "update_id": 4,
                "message": {
                    "chat": {"id": 42, "type": "private"},
                    "from": other,
                    "text": f"/start {self.claim['token']}",
                },
            }
        )
        self.assertIn("другому Telegram-аккаунту", self.api.messages[-1]["text"])

    def test_expired_token_is_rejected(self):
        fixed = datetime(2026, 1, 1, tzinfo=timezone.utc)
        expired = self.database.create_token(
            "expired-session", "TEST-PRIZE", 1, now=fixed
        )
        result, _ = self.database.bind_token(
            expired["token"], self.user, now=fixed + timedelta(hours=2)
        )
        self.assertEqual(result, "expired")

    def test_personal_data_is_encrypted_and_can_be_deleted(self):
        self._start()
        self._consent()
        self._contact()
        raw = self.database.connection.execute(
            "SELECT phone_enc, username_enc FROM prize_tokens"
        ).fetchone()
        self.assertNotIn("79991234567", raw["phone_enc"])
        self.assertNotIn("winner", raw["username_enc"])

        result = self.database.delete_user_data(self.user["id"])
        self.assertEqual(result["claims"], 1)
        deleted = self.database.get_claim(self.claim["token"])
        self.assertEqual(deleted["status"], "cancelled")
        self.assertEqual(deleted["phone"], "")

    def test_regular_user_can_receive_only_one_prize(self):
        self._start()
        self._consent()
        self._contact()
        self.api.member = {"status": "member"}
        self._check("first-prize")
        self.assertEqual(self.database.get_claim(self.claim["token"])["status"], "issued")

        second = self.database.create_token("game-session-2", "TEST-PRIZE", 24)
        self.app.handle_update(
            {
                "update_id": 20,
                "message": {
                    "chat": {"id": self.user["id"], "type": "private"},
                    "from": self.user,
                    "text": f"/start {second['token']}",
                },
            }
        )
        self.assertIn("только один приз", self.api.messages[-1]["text"])
        self.assertIsNone(self.database.get_claim(second["token"])["telegram_user_id"])

    def test_exempt_tester_can_receive_multiple_prizes(self):
        self.user["username"] = "rredactor"
        self._start()
        self._consent()
        self._contact()
        self.api.member = {"status": "member"}
        self._check("tester-first")
        self.assertEqual(self.database.get_claim(self.claim["token"])["status"], "issued")

        self.claim = self.database.create_token("tester-session-2", "TEST-PRIZE", 24)
        self._start()
        self._consent()
        self._contact()
        self._check("tester-second")
        second = self.database.get_claim(self.claim["token"])
        self.assertEqual(second["status"], "issued")
        self.assertEqual(second["promo_code"], "FW-TEST-0002")

    def test_issue_transaction_rechecks_one_prize_limit(self):
        result, _ = self.database.bind_token(self.claim["token"], self.user)
        self.assertEqual(result, "ok")
        self.database.accept_consent(self.claim["token"], self.user)
        self.database.store_phone(
            self.claim["token"], self.user["id"], "+79991234567"
        )
        self.database.mark_subscription(self.claim["token"], self.user["id"])
        first = self.database.reserve_promo(self.claim["token"], self.user["id"])
        issued_at = datetime.now(timezone.utc)
        self.workbook.record_issue(first, issued_at)
        self.database.finalize_issue(self.claim["token"], issued_at)

        second = self.database.create_token("prebound-session-2", "TEST-PRIZE", 24)
        result, _ = self.database.bind_token(
            second["token"], self.user, allow_multiple_prizes=True
        )
        self.assertEqual(result, "ok")
        self.database.accept_consent(second["token"], self.user)
        self.database.store_phone(second["token"], self.user["id"], "+79991234567")
        self.database.mark_subscription(second["token"], self.user["id"])
        with self.assertRaises(PrizeLimitReached):
            self.database.reserve_promo(second["token"], self.user["id"])


class RealWorkbookCompatibilityTest(unittest.TestCase):
    def test_real_workbook_can_be_issued_on_a_temporary_copy(self):
        project_root = Path(__file__).resolve().parents[2]
        source = project_root / "Flexy_Way_промокоды.xlsx"
        if not source.exists():
            self.skipTest("Рабочая книга проекта отсутствует")

        with tempfile.TemporaryDirectory() as temp_name:
            temp_root = Path(temp_name)
            workbook_path = temp_root / source.name
            copy2(source, workbook_path)
            promo_workbook = PromoWorkbook(workbook_path, "Europe/Moscow", "masked")
            database = Database(temp_root / "integration.sqlite3")
            try:
                database.initialize()
                database.sync_inventory(promo_workbook.inventory())
                claim = database.create_token("real-workbook-test", "SUB-L1-04-10", 24)
                user = {
                    "id": 123456789,
                    "username": "integration_test",
                    "first_name": "Тест",
                    "last_name": "Пользователь",
                }
                result, _ = database.bind_token(claim["token"], user)
                self.assertEqual(result, "ok")
                database.accept_consent(claim["token"], user)
                database.store_phone(claim["token"], user["id"], "+79990000000")
                database.mark_subscription(claim["token"], user["id"])
                claim = database.reserve_promo(claim["token"], user["id"])
                issued_at = datetime.now(timezone.utc)
                promo_workbook.record_issue(claim, issued_at)
                finalized = database.finalize_issue(claim["token"], issued_at)
                self.assertEqual(finalized["status"], "issued")

                verification = load_workbook(workbook_path, read_only=False, data_only=False)
                try:
                    promo_sheet = verification["Промокоды"]
                    self.assertEqual(promo_sheet["G5"].value, "Использован")
                    leads = verification["Лиды Telegram"]
                    matching_tokens = [
                        leads.cell(row, 10).value for row in range(5, leads.max_row + 1)
                    ]
                    self.assertIn(claim["token_ref"], matching_tokens)
                finally:
                    verification.close()
            finally:
                database.close()


if __name__ == "__main__":
    unittest.main()
