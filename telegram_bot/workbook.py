from __future__ import annotations

import os
import tempfile
from copy import copy
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Mapping
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.properties import CalcProperties

from .security import mask_phone


PROMO_SHEET = "Промокоды"
LEADS_SHEET = "Лиды Telegram"
HEADER_ROW = 4
LEADS_HEADERS = [
    "Дата и время",
    "Telegram ref",
    "Телефон (маска)",
    "Username",
    "Имя",
    "Фамилия",
    "Код приза",
    "Приз",
    "Промокод",
    "Токен ref",
    "Сессия ref",
    "Статус",
    "Примечание",
]
LEADS_HEADER_ALIASES = {
    "Telegram ID": "Telegram ref",
    "Телефон": "Телефон (маска)",
    "Токен сессии": "Токен ref",
    "ID игровой сессии": "Сессия ref",
}


class WorkbookError(RuntimeError):
    pass


class WorkbookConflict(WorkbookError):
    pass


@dataclass(frozen=True)
class InventoryRecord:
    promo_code: str
    prize_code: str
    prize_name: str
    prize_terms: str
    status: str
    excel_row: int


def _headers(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value is not None
    }


def _excel_safe(value):
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


class PromoWorkbook:
    def __init__(
        self,
        path: Path,
        timezone_name: str = "Europe/Moscow",
        pii_mode: str = "masked",
    ):
        self.path = Path(path)
        self.timezone = ZoneInfo(timezone_name)
        if pii_mode not in {"masked", "full"}:
            raise WorkbookError("Режим персональных данных Excel должен быть masked или full")
        self.pii_mode = pii_mode

    def _load(self, *, read_only: bool = False):
        if not self.path.exists():
            raise WorkbookError(f"Файл промокодов не найден: {self.path}")
        try:
            return load_workbook(self.path, read_only=read_only, data_only=False)
        except Exception as error:
            raise WorkbookError(f"Не удалось открыть файл промокодов: {error}") from error

    def inventory(self) -> list[dict]:
        # Книги, созданные некоторыми генераторами XLSX, не записывают worksheet
        # dimension. В read_only-режиме openpyxl тогда возвращает max_row=None.
        workbook = self._load(read_only=False)
        try:
            if PROMO_SHEET not in workbook.sheetnames:
                raise WorkbookError(f"В книге нет листа «{PROMO_SHEET}»")
            sheet = workbook[PROMO_SHEET]
            columns = _headers(sheet)
            required = {
                "Код приза",
                "Приз",
                "Условия применения",
                "Промокод",
                "Статус",
            }
            missing = required.difference(columns)
            if missing:
                raise WorkbookError("В листе промокодов нет колонок: " + ", ".join(sorted(missing)))

            records = []
            for row in range(HEADER_ROW + 1, sheet.max_row + 1):
                promo_code = sheet.cell(row, columns["Промокод"]).value
                if not promo_code:
                    continue
                record = InventoryRecord(
                    promo_code=str(promo_code).strip(),
                    prize_code=str(sheet.cell(row, columns["Код приза"]).value or "").strip(),
                    prize_name=str(sheet.cell(row, columns["Приз"]).value or "").strip(),
                    prize_terms=str(sheet.cell(row, columns["Условия применения"]).value or "").strip(),
                    status=str(sheet.cell(row, columns["Статус"]).value or "").strip(),
                    excel_row=row,
                )
                records.append(asdict(record))
            if not records:
                raise WorkbookError("В книге не найдено ни одного промокода")
            return records
        finally:
            workbook.close()

    @staticmethod
    def _ensure_leads_sheet(workbook: Workbook):
        if LEADS_SHEET in workbook.sheetnames:
            sheet = workbook[LEADS_SHEET]
            columns = _headers(sheet)
            for old_title, new_title in LEADS_HEADER_ALIASES.items():
                if old_title in columns and new_title not in columns:
                    sheet.cell(HEADER_ROW, columns[old_title], new_title)
            return sheet
        sheet = workbook.create_sheet(LEADS_SHEET)
        sheet.merge_cells("A1:M1")
        sheet["A1"] = "Лиды и выдачи через Telegram-бота"
        sheet.merge_cells("A2:M2")
        sheet["A2"] = (
            "Контакты победителей (имя, фамилия, username, телефон) записываются "
            "для прямой связи менеджера и выдачи призов."
        )
        for column, title in enumerate(LEADS_HEADERS, 1):
            cell = sheet.cell(HEADER_ROW, column, title)
            cell.fill = PatternFill("solid", fgColor="7C3AED")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "C5"
        return sheet

    @staticmethod
    def _lead_row(sheet, token_ref: str) -> tuple[int, bool]:
        token_column = LEADS_HEADERS.index("Токен ref") + 1
        upper = max(sheet.max_row, HEADER_ROW + 1)
        first_empty = None
        for row in range(HEADER_ROW + 1, upper + 1):
            current = sheet.cell(row, token_column).value
            if current == token_ref:
                return row, True
            if first_empty is None and all(
                sheet.cell(row, column).value in (None, "")
                for column in range(1, len(LEADS_HEADERS) + 1)
            ):
                first_empty = row
        return (first_empty or upper + 1), False

    @staticmethod
    def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
        if source_row == target_row:
            return
        for column in range(1, len(LEADS_HEADERS) + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height

    def record_issue(self, claim: Mapping, issued_at: datetime) -> None:
        workbook = self._load(read_only=False)
        temp_path: Path | None = None
        try:
            promo_sheet = workbook[PROMO_SHEET]
            promo_columns = _headers(promo_sheet)
            leads_sheet = self._ensure_leads_sheet(workbook)
            leads_columns = _headers(leads_sheet)

            promo_row = None
            for row in range(HEADER_ROW + 1, promo_sheet.max_row + 1):
                if promo_sheet.cell(row, promo_columns["Промокод"]).value == claim["promo_code"]:
                    promo_row = row
                    break
            if promo_row is None:
                raise WorkbookConflict(f"Промокод {claim['promo_code']} отсутствует в Excel")
            if promo_sheet.cell(promo_row, promo_columns["Код приза"]).value != claim["prize_code"]:
                raise WorkbookConflict("Промокод относится к другому призу")

            lead_row, lead_exists = self._lead_row(leads_sheet, claim["token_ref"])
            current_status = str(
                promo_sheet.cell(promo_row, promo_columns["Статус"]).value or ""
            ).strip()
            if current_status == "Использован" and not lead_exists:
                raise WorkbookConflict("Промокод уже отмечен использованным другим получателем")
            if current_status not in {"Не использован", "Использован"}:
                raise WorkbookConflict(f"Недопустимый статус промокода: {current_status}")

            local_time = issued_at.astimezone(self.timezone).replace(tzinfo=None)
            username = f"@{claim['username']}" if claim.get("username") else ""
            full_name = " ".join(
                part for part in [claim.get("first_name", ""), claim.get("last_name", "")] if part
            )
            if self.pii_mode == "masked_strict":
                recipient = ", ".join(
                    part
                    for part in [claim.get("telegram_user_ref", ""), mask_phone(claim.get("phone"))]
                    if part
                )
                lead_phone = mask_phone(claim.get("phone"))
                lead_username = ""
                lead_first_name = ""
                lead_last_name = ""
            else:
                recipient = ", ".join(
                    part for part in [full_name, username, claim.get("phone", "")] if part
                ) or claim.get("telegram_user_ref", "")
                lead_phone = claim.get("phone", "")
                lead_username = username
                lead_first_name = claim.get("first_name", "")
                lead_last_name = claim.get("last_name", "")

            promo_sheet.cell(promo_row, promo_columns["Статус"], "Использован")
            promo_sheet.cell(promo_row, promo_columns["Дата использования"], local_time)
            promo_sheet.cell(promo_row, promo_columns["Дата использования"]).number_format = "dd.mm.yyyy hh:mm"
            promo_sheet.cell(
                promo_row, promo_columns["Получатель / клиент"], _excel_safe(recipient)
            )
            promo_sheet.cell(
                promo_row,
                promo_columns["Комментарий"],
                f"Выдан Telegram-ботом; получатель: {claim['telegram_user_ref']}; "
                f"сессия: {claim['session_ref']}",
            )

            self._copy_row_style(leads_sheet, HEADER_ROW + 1, lead_row)
            values = {
                "Дата и время": local_time,
                "Telegram ref": claim["telegram_user_ref"],
                "Телефон (маска)": lead_phone,
                "Username": lead_username,
                "Имя": lead_first_name,
                "Фамилия": lead_last_name,
                "Код приза": claim["prize_code"],
                "Приз": claim["prize_name"],
                "Промокод": claim["promo_code"],
                "Токен ref": claim["token_ref"],
                "Сессия ref": claim["session_ref"],
                "Статус": "Выдан",
                "Примечание": "Подписка подтверждена одним запросом Telegram API",
            }
            for title, value in values.items():
                leads_sheet.cell(lead_row, leads_columns[title], _excel_safe(value))
            leads_sheet.cell(lead_row, leads_columns["Дата и время"]).number_format = "dd.mm.yyyy hh:mm"
            leads_sheet.cell(lead_row, leads_columns["Telegram ref"]).number_format = "@"
            leads_sheet.cell(lead_row, leads_columns["Телефон (маска)"]).number_format = "@"
            status_cell = leads_sheet.cell(lead_row, leads_columns["Статус"])
            status_cell.fill = PatternFill("solid", fgColor="ECFDF5")
            status_cell.font = copy(status_cell.font)
            status_cell.font = Font(
                name=status_cell.font.name,
                size=status_cell.font.sz,
                bold=status_cell.font.bold,
                italic=status_cell.font.italic,
                color="166534",
            )

            if workbook.calculation is None:
                workbook.calculation = CalcProperties()
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            file_descriptor, temp_name = tempfile.mkstemp(
                prefix=f".{self.path.stem}.", suffix=".xlsx", dir=self.path.parent
            )
            os.close(file_descriptor)
            temp_path = Path(temp_name)
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, self.path)
            temp_path = None

            verification = load_workbook(self.path, read_only=True, data_only=False)
            try:
                verified_status = verification[PROMO_SHEET].cell(
                    promo_row, promo_columns["Статус"]
                ).value
                verified_token = verification[LEADS_SHEET].cell(
                    lead_row, leads_columns["Токен ref"]
                ).value
                if verified_status != "Использован" or verified_token != claim["token_ref"]:
                    raise WorkbookError("Проверка записи Excel после сохранения не пройдена")
            finally:
                verification.close()
        except PermissionError as error:
            raise WorkbookError(
                "Excel-файл занят. Закройте его в Microsoft Excel и повторите проверку подписки."
            ) from error
        except WorkbookError:
            raise
        except Exception as error:
            raise WorkbookError(f"Не удалось обновить Excel: {error}") from error
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            if temp_path is not None and temp_path.exists():
                temp_path.unlink()

    def anonymize_user(self, user_ref: str) -> int:
        return self.anonymize_users({user_ref})

    def anonymize_users(self, user_refs: set[str] | list[str]) -> int:
        user_refs = set(user_refs)
        if not user_refs:
            return 0
        workbook = self._load(read_only=False)
        temp_path: Path | None = None
        changed = 0
        try:
            if LEADS_SHEET not in workbook.sheetnames:
                return 0
            leads_sheet = self._ensure_leads_sheet(workbook)
            columns = _headers(leads_sheet)
            for row in range(HEADER_ROW + 1, leads_sheet.max_row + 1):
                if leads_sheet.cell(row, columns["Telegram ref"]).value not in user_refs:
                    continue
                for title in ("Telegram ref", "Телефон (маска)", "Username", "Имя", "Фамилия"):
                    leads_sheet.cell(row, columns[title], "Удалено по запросу")
                changed += 1
            if not changed:
                return 0

            file_descriptor, temp_name = tempfile.mkstemp(
                prefix=f".{self.path.stem}.", suffix=".xlsx", dir=self.path.parent
            )
            os.close(file_descriptor)
            temp_path = Path(temp_name)
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, self.path)
            temp_path = None
            return changed
        except PermissionError as error:
            raise WorkbookError(
                "Excel-файл занят. Закройте его в Microsoft Excel и повторите удаление данных."
            ) from error
        except WorkbookError:
            raise
        except Exception as error:
            raise WorkbookError(f"Не удалось обезличить данные в Excel: {error}") from error
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            if temp_path is not None and temp_path.exists():
                temp_path.unlink()
